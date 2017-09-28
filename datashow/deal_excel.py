import xlrd
import xlwt
import struct
import openpyxl
from openpyxl import Workbook

from openpyxl import load_workbook

from openpyxl.writer.excel import ExcelWriter

def read_excel():
    workbook = xlrd.open_workbook(filename='data/account.xls', on_demand=True)
    worksheet = workbook.sheet_by_index(0)
    data = []

    for r in xrange(worksheet.nrows):
        col = []
        for c in range(worksheet.ncols):
            col.append(worksheet.cell(r, c).value)
        data.append(col)
    return data

def write_new_excel():
    data = read_excel()
    newbook = xlwt.Workbook()
    wsheet = newbook.add_sheet("added")
    r = 0
    for line in data:
        c = 0
        for d in line:
            wsheet.write(r, c, d)
            c += 1
        r += 1

    newbook.save('data/account1.xls')

def struct_read():
    with open('data/data.csv') as f:
        for line in f:
            fields = struct.Struct('2s1s4s').unpack_from(line)
            print "fields", [field.strip() for field in fields]

def add_excel_sheet():
    workbook_ = load_workbook('data/account.xlsx')
    newsheet = workbook_.create_sheet()
    newsheet.append([1,2,3,4])
    workbook_.copy_worksheet(workbook_.get_sheet_by_name('added'))
    workbook_.save('data/account.xlsx')

add_excel_sheet()



